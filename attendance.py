import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import sys
from collections import OrderedDict

class StudentRegistry:
    """Handles reading and caching student information from Excel file."""
    
    def __init__(self, filename='studentinfo.xlsx'):
        self.filename = filename
        self.student_cache = {}  # {student_id: rfid}
        self.rfid_cache = {}     # {rfid: student_id}
        self.student_names = {}  # {student_id: name} for display
        self.load_students()
    
    def load_students(self):
        """Load student data from Excel file into memory for fast lookup."""
        try:
            if not os.path.exists(self.filename):
                return False, f"Error: {self.filename} not found!"
            
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            
            # Column A = Student ID, Column K = RFID, Column B = Name (optional)
            for row in ws.iter_rows(min_row=2, values_only=True):
                student_id = row[0]  # Column A
                name = row[1] if row[1] else ""  # Column B (optional)
                rfid = row[10]       # Column K
                
                if student_id and rfid:
                    self.student_cache[str(student_id).strip()] = str(rfid).strip()
                    self.rfid_cache[str(rfid).strip()] = str(student_id).strip()
                    self.student_names[str(student_id).strip()] = str(name).strip()
            
            return True, f"Loaded {len(self.student_cache)} students"
        except Exception as e:
            return False, f"Error loading student data: {e}"
    
    def get_student_by_id(self, student_id):
        """Get RFID by student ID."""
        return self.student_cache.get(str(student_id).strip())
    
    def get_student_by_rfid(self, rfid):
        """Get student ID by RFID."""
        return self.rfid_cache.get(str(rfid).strip())
    
    def get_student_name(self, student_id):
        """Get student name if available."""
        return self.student_names.get(str(student_id).strip(), "")
    
    def is_valid_student(self, identifier):
        """Check if identifier (student ID or RFID) exists."""
        identifier = str(identifier).strip()
        return identifier in self.student_cache or identifier in self.rfid_cache
    
    def lookup_student(self, identifier):
        """Lookup student by ID or RFID, return (student_id, rfid, status)."""
        identifier = str(identifier).strip()
        
        # Check if it's a RFID
        if identifier in self.rfid_cache:
            student_id = self.rfid_cache[identifier]
            return student_id, identifier, "FOUND_BY_RFID"
        
        # Check if it's a student ID
        if identifier in self.student_cache:
            rfid = self.student_cache[identifier]
            return identifier, rfid, "FOUND_BY_ID"
        
        # Not found - return as-is with UNKNOWN status
        return identifier, identifier, "UNKNOWN"


class AttendanceSheet:
    """Handles attendance recording and Excel export."""
    
    def __init__(self, filename, staff_name, staff_number, event_name, venue, start_time, end_time):
        self.filename = filename
        self.staff_name = staff_name
        self.staff_number = staff_number
        self.event_name = event_name
        self.venue = venue
        self.start_time = start_time
        self.end_time = end_time
        self.records = OrderedDict()  # {input_id: (timestamp, rfid, student_id, status)} 
        self.all_inputs = []  # All inputs for logging, including unknowns
    
    def add_record(self, input_id, rfid, student_id, status):
        """
        Add an attendance record (saves ALL inputs regardless of validity).
        status: "FOUND_BY_RFID", "FOUND_BY_ID", or "UNKNOWN"
        """
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Log all inputs
        self.all_inputs.append({
            'timestamp': timestamp,
            'input': input_id,
            'rfid': rfid,
            'student_id': student_id,
            'status': status
        })
        
        # Only add to primary records if valid student
        if status != "UNKNOWN":
            if student_id not in self.records:
                self.records[student_id] = (timestamp, rfid, student_id, status)
                return timestamp, True, "Valid"
            return self.records[student_id][0], False, "Duplicate"
        else:
            return timestamp, False, "Unknown ID/RFID"
    
    def save_to_excel(self):
        """Save attendance records to Excel file with multiple sheets."""
        try:
            wb = openpyxl.Workbook()
            
            # ===== SHEET 1: Attendance (Valid entries only) =====
            ws_attendance = wb.active
            ws_attendance.title = "Attendance"
            
            # Add header information
            ws_attendance['A1'] = "Event Details"
            ws_attendance['A1'].font = Font(bold=True, size=12)
            ws_attendance['A2'] = f"Staff Name: {self.staff_name}"
            ws_attendance['A3'] = f"Staff Number: {self.staff_number}"
            ws_attendance['A4'] = f"Event: {self.event_name}"
            ws_attendance['A5'] = f"Venue: {self.venue}"
            ws_attendance['A6'] = f"Start Time: {self.start_time}"
            ws_attendance['A7'] = f"End Time: {self.end_time}"
            ws_attendance['A8'] = f"Total Attendees: {len(self.records)}"
            
            # Column headers
            header_row = 10
            ws_attendance[f'A{header_row}'] = "Time Scanned"
            ws_attendance[f'B{header_row}'] = "RFID"
            ws_attendance[f'C{header_row}'] = "Student ID"
            
            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for col in ['A', 'B', 'C']:
                ws_attendance[f'{col}{header_row}'].fill = header_fill
                ws_attendance[f'{col}{header_row}'].font = header_font
                ws_attendance[f'{col}{header_row}'].alignment = Alignment(horizontal="center")
            
            # Add valid records
            for idx, (student_id, (timestamp, rfid, _, _)) in enumerate(self.records.items(), start=1):
                row = header_row + idx
                ws_attendance[f'A{row}'] = timestamp
                ws_attendance[f'B{row}'] = rfid
                ws_attendance[f'C{row}'] = student_id
            
            # Adjust column widths
            ws_attendance.column_dimensions['A'].width = 20
            ws_attendance.column_dimensions['B'].width = 25
            ws_attendance.column_dimensions['C'].width = 15
            
            # ===== SHEET 2: Complete Log (All inputs) =====
            ws_log = wb.create_sheet("Complete Log")
            
            log_header_row = 1
            ws_log[f'A{log_header_row}'] = "Timestamp"
            ws_log[f'B{log_header_row}'] = "Input Scanned"
            ws_log[f'C{log_header_row}'] = "RFID"
            ws_log[f'D{log_header_row}'] = "Student ID"
            ws_log[f'E{log_header_row}'] = "Status"
            
            # Style log headers
            log_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            log_font = Font(bold=True, color="FFFFFF")
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws_log[f'{col}{log_header_row}'].fill = log_fill
                ws_log[f'{col}{log_header_row}'].font = log_font
                ws_log[f'{col}{log_header_row}'].alignment = Alignment(horizontal="center")
            
            # Add all inputs to log
            for idx, log_entry in enumerate(self.all_inputs, start=1):
                row = log_header_row + idx
                ws_log[f'A{row}'] = log_entry['timestamp']
                ws_log[f'B{row}'] = log_entry['input']
                ws_log[f'C{row}'] = log_entry['rfid']
                ws_log[f'D{row}'] = log_entry['student_id']
                ws_log[f'E{row}'] = log_entry['status']
                
                # Color code status
                if log_entry['status'] == "UNKNOWN":
                    ws_log[f'E{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif log_entry['status'] == "FOUND_BY_RFID":
                    ws_log[f'E{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    ws_log[f'E{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            # Adjust log column widths
            ws_log.column_dimensions['A'].width = 20
            ws_log.column_dimensions['B'].width = 20
            ws_log.column_dimensions['C'].width = 25
            ws_log.column_dimensions['D'].width = 15
            ws_log.column_dimensions['E'].width = 15
            
            # ===== SHEET 3: Summary =====
            ws_summary = wb.create_sheet("Summary", 0)
            
            ws_summary['A1'] = "ATTENDANCE SUMMARY"
            ws_summary['A1'].font = Font(bold=True, size=14)
            
            ws_summary['A3'] = "Event Information"
            ws_summary['A3'].font = Font(bold=True, size=12)
            ws_summary['A4'] = f"Event Name: {self.event_name}"
            ws_summary['A5'] = f"Venue: {self.venue}"
            ws_summary['A6'] = f"Staff: {self.staff_name} ({self.staff_number})"
            ws_summary['A7'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
            ws_summary['A8'] = f"Time Range: {self.start_time} - {self.end_time}"
            
            ws_summary['A10'] = "Statistics"
            ws_summary['A10'].font = Font(bold=True, size=12)
            
            valid_count = len(self.records)
            unknown_count = sum(1 for log in self.all_inputs if log['status'] == "UNKNOWN")
            total_inputs = len(self.all_inputs)
            
            ws_summary['A11'] = f"Total Inputs: {total_inputs}"
            ws_summary['A12'] = f"Valid Attendees: {valid_count}"
            ws_summary['A13'] = f"Unknown/Invalid IDs: {unknown_count}"
            
            wb.save(self.filename)
            return True, f"Saved {len(self.records)} valid + {len(self.all_inputs)} total log entries to {self.filename}"
        except Exception as e:
            return False, f"Error saving attendance file: {e}"


class AttendanceUI:
    """Modern, fast, responsive UI with minimal human intervention."""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Attendance System - RFID Scanner")
        self.root.geometry("1000x700")
        self.root.configure(bg='#f0f0f0')
        
        self.student_registry = StudentRegistry()
        self.current_sheet = None
        self.scanned_students = []
        
        self.setup_styles()
        self.show_menu()
    
    def setup_styles(self):
        """Setup TTK styles for modern look."""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Colors
        style.configure('Title.TLabel', font=('Helvetica', 24, 'bold'), background='#f0f0f0', foreground='#1a5490')
        style.configure('Header.TLabel', font=('Helvetica', 14, 'bold'), background='#f0f0f0')
        style.configure('Normal.TLabel', font=('Helvetica', 11), background='#f0f0f0')
        style.configure('Success.TLabel', font=('Helvetica', 12, 'bold'), foreground='#28a745', background='#f0f0f0')
        style.configure('Error.TLabel', font=('Helvetica', 12, 'bold'), foreground='#dc3545', background='#f0f0f0')
        style.configure('Warning.TLabel', font=('Helvetica', 12, 'bold'), foreground='#ff9800', background='#f0f0f0')
        style.configure('Info.TLabel', font=('Helvetica', 10), background='#f0f0f0', foreground='#666')
        
        style.configure('Action.TButton', font=('Helvetica', 12, 'bold'), padding=10)
        style.configure('TEntry', font=('Helvetica', 11), padding=5)
    
    def clear_window(self):
        """Clear all widgets from window."""
        for widget in self.root.winfo_children():
            widget.destroy()
    
    def show_menu(self):
        """Show main menu."""
        self.clear_window()
        
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title = ttk.Label(main_frame, text="📋 ATTENDANCE SYSTEM", style='Title.TLabel')
        title.pack(pady=20)
        
        # Status label showing loaded students
        status_ok, status_msg = self.student_registry.load_students()
        status_color = "Success.TLabel" if status_ok else "Error.TLabel"
        status = ttk.Label(main_frame, text=f"✓ {status_msg}" if status_ok else f"✗ {status_msg}", style=status_color)
        status.pack(pady=10)
        
        # Button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=40)
        
        btn_new = ttk.Button(
            btn_frame, 
            text="➕ CREATE NEW ATTENDANCE", 
            command=self.show_create_sheet,
            style='Action.TButton',
            width=30
        )
        btn_new.pack(pady=10)
        
        btn_load = ttk.Button(
            btn_frame, 
            text="📂 LOAD EXISTING ATTENDANCE", 
            command=self.show_load_sheet,
            style='Action.TButton',
            width=30
        )
        btn_load.pack(pady=10)
        
        btn_quit = ttk.Button(
            btn_frame, 
            text="❌ EXIT", 
            command=self.root.quit,
            style='Action.TButton',
            width=30
        )
        btn_quit.pack(pady=10)
    
    def show_create_sheet(self):
        """Show form to create new attendance sheet."""
        self.clear_window()
        
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title = ttk.Label(main_frame, text="📝 NEW ATTENDANCE SHEET", style='Title.TLabel')
        title.pack(pady=20)
        
        # Form frame
        form_frame = ttk.Frame(main_frame)
        form_frame.pack(fill=tk.X, pady=20)
        
        fields = {}
        field_names = ["Staff Name", "Staff Student Number", "Event Name", "Venue", "Start Time (HH:MM)", "End Time (HH:MM)"]
        
        for field_name in field_names:
            row = ttk.Frame(form_frame)
            row.pack(fill=tk.X, pady=8)
            
            label = ttk.Label(row, text=f"{field_name}:", style='Normal.TLabel', width=25)
            label.pack(side=tk.LEFT)
            
            entry = ttk.Entry(row, width=40)
            entry.pack(side=tk.LEFT, padx=10)
            fields[field_name] = entry
        
        # Button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)
        
        def create():
            values = {k: v.get().strip() for k, v in fields.items()}
            
            if not all(values.values()):
                messagebox.showerror("Error", "All fields are required!")
                return
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"attendance_{values['Event Name'].replace(' ', '_')}_{timestamp}.xlsx"
            
            self.current_sheet = AttendanceSheet(
                filename,
                values['Staff Name'],
                values['Staff Student Number'],
                values['Event Name'],
                values['Venue'],
                values['Start Time (HH:MM)'],
                values['End Time (HH:MM)']
            )
            
            self.show_scanning_interface()
        
        btn_create = ttk.Button(btn_frame, text="✓ START SCANNING", command=create, style='Action.TButton', width=30)
        btn_create.pack(side=tk.LEFT, padx=10)
        
        btn_back = ttk.Button(btn_frame, text="← BACK", command=self.show_menu, style='Action.TButton', width=30)
        btn_back.pack(side=tk.LEFT, padx=10)
    
    def show_loading_interface(self, filename):
        """Show scanning interface for loaded attendance sheet."""
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb['Summary'] if 'Summary' in wb.sheetnames else wb.active
            
            # Extract event info from header
            staff_name = str(ws['A4'].value or "").replace("Staff Name: ", "")
            staff_number = str(ws['A6'].value or "").replace("Staff: ", "").split(" (")[0]
            event_name = str(ws['A3'].value or "").replace("Event Name: ", "")
            venue = str(ws['A5'].value or "").replace("Venue: ", "")
            start_time = str(ws['A8'].value or "").split(" - ")[0].replace("Time Range: ", "")
            end_time = str(ws['A8'].value or "").split(" - ")[1] if " - " in str(ws['A8'].value or "") else ""
            
            self.current_sheet = AttendanceSheet(filename, staff_name, staff_number, event_name, venue, start_time, end_time)
            
            # Load existing records
            if 'Attendance' in wb.sheetnames:
                ws_att = wb['Attendance']
                for row in ws_att.iter_rows(min_row=11, values_only=True):
                    if row[0] and row[1] and row[2]:  # timestamp, rfid, student_id
                        self.current_sheet.records[row[2]] = (row[0], row[1], row[2], "FOUND_BY_ID")
            
            # Load all log entries
            if 'Complete Log' in wb.sheetnames:
                ws_log = wb['Complete Log']
                for row in ws_log.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # timestamp exists
                        self.current_sheet.all_inputs.append({
                            'timestamp': row[0],
                            'input': row[1],
                            'rfid': row[2],
                            'student_id': row[3],
                            'status': row[4]
                        })
            
            self.show_scanning_interface()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file: {e}")
            self.show_load_sheet()
    
    def show_load_sheet(self):
        """Show list of existing attendance sheets."""
        self.clear_window()
        
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        title = ttk.Label(main_frame, text="📂 LOAD ATTENDANCE", style='Title.TLabel')
        title.pack(pady=20)
        
        # Find attendance files
        attendance_files = sorted([f for f in os.listdir('.') if f.startswith('attendance_') and f.endswith('.xlsx')], reverse=True)
        
        if not attendance_files:
            msg = ttk.Label(main_frame, text="No attendance files found.", style='Info.TLabel')
            msg.pack(pady=20)
            
            btn_back = ttk.Button(main_frame, text="← BACK", command=self.show_menu, style='Action.TButton', width=30)
            btn_back.pack(pady=20)
            return
        
        # Listbox with files
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, font=('Helvetica', 11), height=15)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)
        
        for file in attendance_files:
            listbox.insert(tk.END, file)
        
        def load_selected():
            selection = listbox.curselection()
            if selection:
                filename = attendance_files[selection[0]]
                self.show_loading_interface(filename)
            else:
                messagebox.showwarning("Warning", "Please select a file!")
        
        # Button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)
        
        btn_load = ttk.Button(btn_frame, text="✓ LOAD", command=load_selected, style='Action.TButton', width=30)
        btn_load.pack(side=tk.LEFT, padx=10)
        
        btn_back = ttk.Button(btn_frame, text="← BACK", command=self.show_menu, style='Action.TButton', width=30)
        btn_back.pack(side=tk.LEFT, padx=10)
    
    def show_scanning_interface(self):
        """Show fast, responsive scanning interface."""
        self.clear_window()
        
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Header with event info
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=10, padx=10)
        
        event_info = f"{self.current_sheet.event_name} | {self.current_sheet.venue}"
        title = ttk.Label(header_frame, text=f"🎯 {event_info}", style='Header.TLabel')
        title.pack()
        
        # Scanning input - FOCUSED FOR INSTANT INPUT
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=10, padx=10)
        
        ttk.Label(input_frame, text="Scan RFID or Enter Student ID:", style='Normal.TLabel').pack()
        
        self.scan_input = ttk.Entry(input_frame, font=('Helvetica', 14), width=40)
        self.scan_input.pack(pady=5)
        self.scan_input.focus()
        
        # Status label for feedback
        self.status_label = ttk.Label(main_frame, text="Ready to scan...", style='Info.TLabel')
        self.status_label.pack(pady=5)
        
        # Scanned list with scrollbar
        list_frame = ttk.LabelFrame(main_frame, text=f"Scanned: {len(self.current_sheet.records)} Valid | {len(self.current_sheet.all_inputs)} Total", padding=5)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10, padx=10)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.scan_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, font=('Courier', 9), height=15)
        self.scan_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.scan_listbox.yview)
        
        self.populate_scan_list()
        
        # Button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10, padx=10)
        
        btn_save = ttk.Button(btn_frame, text="💾 SAVE & FINISH", command=self.finish_scanning, style='Action.TButton')
        btn_save.pack(side=tk.LEFT, padx=5)
        
        btn_clear = ttk.Button(btn_frame, text="🔄 CLEAR ALL", command=self.clear_all_records, style='Action.TButton')
        btn_clear.pack(side=tk.LEFT, padx=5)
        
        btn_back = ttk.Button(btn_frame, text="← BACK", command=self.show_menu, style='Action.TButton')
        btn_back.pack(side=tk.RIGHT, padx=5)
        
        # Bind Enter key for instant scanning
        self.scan_input.bind('<Return>', self.process_scan)
    
    def process_scan(self, event=None):
        """Process scanned RFID or student ID - instant response. SAVES ALL INPUTS."""
        identifier = self.scan_input.get().strip()
        self.scan_input.delete(0, tk.END)
        self.scan_input.focus()
        
        if not identifier:
            return
        
        # Lookup student (returns even if UNKNOWN)
        student_id, rfid, status = self.student_registry.lookup_student(identifier)
        
        # Add record (ALWAYS saves regardless of validity)
        timestamp, is_new, message = self.current_sheet.add_record(identifier, rfid, student_id, status)
        
        # Update UI based on result
        if status == "UNKNOWN":
            self.status_label.config(text=f"⚠ UNKNOWN ID: {identifier} (Logged)", style='Warning.TLabel')
            self.scan_input.config(foreground='orange')
        elif is_new:
            student_name = self.student_registry.get_student_name(student_id)
            name_display = f" ({student_name})" if student_name else ""
            self.status_label.config(text=f"✓ {student_id}{name_display} - {timestamp}", style='Success.TLabel')
            self.scan_input.config(foreground='green')
        else:
            self.status_label.config(text=f"⚠ DUPLICATE: {student_id} (Logged)", style='Warning.TLabel')
            self.scan_input.config(foreground='orange')
        
        self.populate_scan_list()
        self.root.after(1500, lambda: self.scan_input.config(foreground='black'))
    
    def populate_scan_list(self):
        """Update the scanned list display."""
        self.scan_listbox.delete(0, tk.END)
        valid_count = len(self.current_sheet.records)
        total_count = len(self.current_sheet.all_inputs)
        
        self.scan_listbox.insert(tk.END, f"{'#':<3} {'Student ID':<15} {'RFID':<25} {'Timestamp':<20} {'Status':<10}")
        self.scan_listbox.insert(tk.END, "-" * 75)
        
        for idx, log_entry in enumerate(self.current_sheet.all_inputs, 1):
            status_icon = "✓" if log_entry['status'] != "UNKNOWN" else "✗"
            display = f"{idx:<3} {log_entry['student_id']:<15} {log_entry['rfid']:<25} {log_entry['timestamp']:<20} {status_icon} {log_entry['status']}"
            self.scan_listbox.insert(tk.END, display)
        
        # Scroll to bottom
        self.scan_listbox.see(tk.END)
    
    def clear_all_records(self):
        """Clear all scanned records."""
        if messagebox.askyesno("Confirm", "Clear all scanned records?"):
            self.current_sheet.records.clear()
            self.current_sheet.all_inputs.clear()
            self.populate_scan_list()
            self.status_label.config(text="Records cleared", style='Info.TLabel')
    
    def finish_scanning(self):
        """Save attendance and return to menu."""
        if not self.current_sheet.all_inputs:
            messagebox.showwarning("Warning", "No records to save!")
            return
        
        success, message = self.current_sheet.save_to_excel()
        
        if success:
            messagebox.showinfo("Success", message)
            self.current_sheet = None
            self.show_menu()
        else:
            messagebox.showerror("Error", message)


def main():
    """Entry point."""
    root = tk.Tk()
    app = AttendanceUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
